import os

from .consts import *
import csv
import openpyxl
import json


class LHSProject:
    def init_dirs(self):
        os.makedirs(PRE_PARATRANZ_RESULTS_DIR, exist_ok=True)
        os.makedirs(POST_PARATRANZ_RESULTS_DIR, exist_ok=True)

    def process_csv_file_to_paratranz_json(self):
        """csv => json"""
        for file in os.listdir(RAW_FILES_DIR):
            filename = file.split(".")[0]

            if file.endswith(".xlsx"):
                wb = openpyxl.load_workbook(RAW_FILES_DIR / file)
                ws = wb["Sheet1"]
                raw_data = list(ws.values)
            else:
                with open(RAW_FILES_DIR / file, "r", encoding="utf-8") as fp:
                    raw_data = list(csv.reader(fp))

            json_data_en = [
                {
                    "key": item[0],
                    "original": item[1],
                    "translation": ""
                } for item in raw_data
                if item[1]
            ]

            json_data_sp = [
                {
                    "key": item[0],
                    "original": item[2],
                    "translation": ""
                } for item in raw_data
                if item[2]
            ]

            with open(PRE_PARATRANZ_RESULTS_DIR / f"{filename}-en.json", "w", encoding="utf-8") as fp:
                json.dump(json_data_en, fp, ensure_ascii=False, indent=2)
            with open(PRE_PARATRANZ_RESULTS_DIR / f"{filename}-sp.json", "w", encoding="utf-8") as fp:
                json.dump(json_data_sp, fp, ensure_ascii=False, indent=2)

    def process_paratranz_json_to_csv_file(self):
        """json => csv"""
        for file in os.listdir(RAW_FILES_DIR):
            filename = file.split(".")[0]

            if file.endswith(".xlsx"):
                wb = openpyxl.load_workbook(RAW_FILES_DIR / file)
                ws = wb["Sheet1"]
                raw_data = list(ws.values)
                raw_data = [
                    list(data)
                    for data in raw_data
                ]
            else:
                with open(RAW_FILES_DIR / file, "r", encoding="utf-8") as fp:
                    raw_data = list(csv.reader(fp))

            with open(DIR_PARATRANZ / "utf8" / f"{filename}-en.json", "r", encoding="utf-8") as fp:
                en_data = json.load(fp)
            with open(DIR_PARATRANZ / "utf8" / f"{filename}-sp.json", "r", encoding="utf-8") as fp:
                sp_data = json.load(fp)

            en_data = {
                item["key"]: item["translation"]
                for item in en_data
                if item["translation"]
            }
            sp_data = {
                item["key"]: item["translation"]
                for item in sp_data
                if item["translation"]
            }

            for idx, line in enumerate(raw_data):
                key = line[0]
                if en_data.get(key):
                    raw_data[idx][1] = en_data[key]
                if sp_data.get(key):
                    raw_data[idx][2] = sp_data[key]

            if file.endswith(".xlsx"):
                wb = openpyxl.Workbook()
                ws = wb.active
                for row in raw_data:
                    ws.append(row)
                wb.save(POST_PARATRANZ_RESULTS_DIR / file)

            else:
                with open(POST_PARATRANZ_RESULTS_DIR / file, "w", encoding="utf-8-sig", newline="") as fp:
                    csv.writer(fp).writerows(raw_data)


__all__ = [
    "LHSProject"
]
